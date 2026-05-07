#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 规则引擎数据集分类.xlsx 导入数据集和数据元到数据库
用法: python3 docs/import_dataset_dataelement.py
"""

import sys
import os
import pymysql
from collections import OrderedDict

def get_db_connection():
    return pymysql.connect(
        host='localhost',
        user='root',
        password='qiubin78',
        database='ruleengine',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

def load_excel_data(path):
    try:
        import openpyxl
    except ImportError:
        print("错误: 需要安装 openpyxl: pip3 install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['数据元']

    headers = [cell.value for cell in ws[2]]
    print(f"Excel 列头: {headers}")

    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        rec = {
            'catL1Code': str(row[0]).strip() if row[0] else '',
            'catL1Name': str(row[1]).strip() if row[1] else '',
            'catL2Code': str(row[2]).strip() if row[2] else '',
            'catL2Name': str(row[3]).strip() if row[3] else '',
            'catL3Code': str(row[4]).strip() if row[4] else '',
            'catL3Name': str(row[5]).strip() if row[5] else '',
            'datasetEnglishName': str(row[6]).strip() if row[6] else '',
            'datasetSortOrder': int(row[7]) if row[7] is not None else 0,
            'standardCode': str(row[8]).strip() if row[8] else '',
            'englishName': str(row[9]).strip() if row[9] else '',
            'camelName': str(row[10]).strip() if row[10] else '',
            'name': str(row[11]).strip() if row[11] else '',
            'definition': str(row[12]).strip() if row[12] else '',
            'sortOrder': int(row[13]) if row[13] is not None else 0,
            'sensitivity': str(row[14]).strip() if row[14] else '',
        }
        # 跳过空行（无数据元名称或驼峰名称）
        if not rec['camelName'] and not rec['name']:
            continue
        records.append(rec)

    return records

def clear_existing_data(conn):
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM condition_model_operators")
        cursor.execute("DELETE FROM condition_model")
        cursor.execute("DELETE FROM data_element")
        cursor.execute("DELETE FROM data_set")
        print("已清空现有 data_set, data_element, condition_model, condition_model_operators")

def ensure_categories(conn, records):
    """确保一级分类存在，返回 {catL1Code: category_id}"""
    categories = {}
    seen = set()
    for rec in records:
        code = rec['catL1Code']
        name = rec['catL1Name']
        if not code or code in seen:
            continue
        seen.add(code)
        categories[code] = {'code': f'CAT_{code}', 'name': name or code, 'sort': int(code) if code.isdigit() else 0}

    with conn.cursor() as cursor:
        # 先清空自动创建的分类
        cursor.execute("SELECT id FROM condition_model_category WHERE code LIKE 'CAT_%'")
        existing = cursor.fetchall()
        for row in existing:
            cat_id = row['id']
            cursor.execute("DELETE FROM condition_model_operators WHERE condition_model_id IN (SELECT id FROM condition_model WHERE category_id=%s)", (cat_id,))
            cursor.execute("DELETE FROM condition_model WHERE category_id=%s", (cat_id,))
        cursor.execute("DELETE FROM condition_model_category WHERE code LIKE 'CAT_%'")

        result = {}
        for code, info in sorted(categories.items(), key=lambda x: x[1]['sort']):
            cursor.execute(
                "INSERT INTO condition_model_category (code, name, description, sort_order, status, created_at, updated_at) VALUES (%s, %s, %s, %s, 'ENABLED', NOW(), NOW())",
                (info['code'], info['name'], f"自动创建自数据元导入，一级分类: {code}", info['sort'])
            )
            result[code] = cursor.lastrowid
        conn.commit()
        print(f"已创建/更新 {len(result)} 个条件模型分类")
        return result

def upsert_datasets(conn, records):
    """插入数据集，返回 {catL3Code: dataset_id}"""
    datasets = OrderedDict()
    for rec in records:
        catL3Code = rec['catL3Code']
        if not catL3Code or catL3Code in datasets:
            continue
        datasets[catL3Code] = rec

    result = {}
    with conn.cursor() as cursor:
        for catL3Code, rec in datasets.items():
            cursor.execute(
                """INSERT INTO data_set (code, name, english_name, cat_l1_code, cat_l1_name, cat_l2_code, cat_l2_name, cat_l3_code, cat_l3_name, sort_order, status, created_at, updated_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ENABLED', NOW(), NOW())""",
                (catL3Code,
                 rec['catL3Name'] or catL3Code,
                 rec['datasetEnglishName'],
                 rec['catL1Code'],
                 rec['catL1Name'],
                 rec['catL2Code'],
                 rec['catL2Name'],
                 catL3Code,
                 rec['catL3Name'],
                 rec['datasetSortOrder'])
            )
            result[catL3Code] = cursor.lastrowid
        conn.commit()
        print(f"已插入 {len(result)} 个数据集")
        return result

def upsert_data_elements(conn, records, dataset_map):
    """插入数据元，返回 [(data_element_id, data_type, camel_name, name)]"""
    elements = []
    seen = set()
    with conn.cursor() as cursor:
        for rec in records:
            catL3Code = rec['catL3Code']
            camelName = rec['camelName']
            if not catL3Code or not camelName:
                continue
            ds_id = dataset_map.get(catL3Code)
            if not ds_id:
                continue
            code = f"{catL3Code}.{camelName}"
            if code in seen:
                continue
            seen.add(code)

            # 默认数据类型 STRING
            data_type = 'STRING'

            cursor.execute(
                """INSERT INTO data_element (code, name, data_type, standard_code, english_name, camel_name, definition, sort_order, sensitivity, dataset_id, status, created_at, updated_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ENABLED', NOW(), NOW())""",
                (code,
                 rec['name'] or camelName,
                 data_type,
                 rec['standardCode'] if rec['standardCode'] else None,
                 rec['englishName'] if rec['englishName'] else None,
                 camelName,
                 rec['definition'] if rec['definition'] else None,
                 rec['sortOrder'],
                 rec['sensitivity'] if rec['sensitivity'] else None,
                 ds_id)
            )
            elements.append({
                'id': cursor.lastrowid,
                'data_type': data_type,
                'camel_name': camelName,
                'name': rec['name'] or camelName,
                'code': code
            })
        conn.commit()
        print(f"已插入 {len(elements)} 个数据元")
        return elements

def default_operators(data_type):
    if data_type == 'STRING':
        return ["==", "contains", "regexMatch", "lengthCheck", "isBlank", "similarity"]
    elif data_type == 'NUMERIC':
        return ["==", "!=", ">", "<", ">=", "<=", "dataCheck"]
    elif data_type == 'DICTIONARY':
        return ["=="]
    elif data_type == 'DICTIONARY_SET':
        return ["IN_SET"]
    elif data_type == 'DATE_TIME':
        return ["timeCheck"]
    elif data_type == 'BOOLEAN':
        return ["==", "isTrue", "isFalse"]
    elif data_type == 'STRING_LIST':
        return ["arrayLength", "arrayIntersect"]
    else:
        return ["=="]

def auto_create_condition_models(conn, elements, category_map, records):
    """为每个数据元自动创建条件模型"""
    # 建立 catL3Code -> catL1Code 映射
    ds_to_cat = {}
    for rec in records:
        ds_to_cat[rec['catL3Code']] = rec['catL1Code']

    with conn.cursor() as cursor:
        count = 0
        for elem in elements:
            # 解析 code 得到 catL3Code
            parts = elem['code'].split('.')
            catL3Code = parts[0] if parts else ''
            catL1Code = ds_to_cat.get(catL3Code)
            category_id = category_map.get(catL1Code) if catL1Code else None

            operators = default_operators(elem['data_type'])
            ops_json = json.dumps(operators)

            cursor.execute(
                """INSERT INTO condition_model (code, name, data_type, value_source, data_element_id, category_id, node_usage, created_at, updated_at)
                   VALUES (%s, %s, %s, 'ADAPTER', %s, %s, 'CONDITION', NOW(), NOW())""",
                (elem['camel_name'], elem['name'], elem['data_type'], elem['id'], category_id)
            )
            cm_id = cursor.lastrowid

            for op in operators:
                cursor.execute(
                    "INSERT INTO condition_model_operators (condition_model_id, operator) VALUES (%s, %s)",
                    (cm_id, op)
                )
            count += 1
        conn.commit()
        print(f"已自动创建 {count} 个条件模型")

def main():
    excel_path = 'docs/规则引擎数据集分类.xlsx'
    if not os.path.exists(excel_path):
        print(f"错误: 找不到Excel文件: {excel_path}")
        sys.exit(1)

    print("正在读取 Excel...")
    records = load_excel_data(excel_path)
    print(f"共读取 {len(records)} 条数据元记录")

    # 统计信息
    ds_count = len(set(r['catL3Code'] for r in records if r['catL3Code']))
    cat_count = len(set(r['catL1Code'] for r in records if r['catL1Code']))
    print(f"涉及 {cat_count} 个一级分类, {ds_count} 个数据集")

    print("\n连接数据库...")
    conn = get_db_connection()

    try:
        print("清空现有数据...")
        clear_existing_data(conn)

        print("\n导入条件模型分类...")
        category_map = ensure_categories(conn, records)

        print("\n导入数据集...")
        dataset_map = upsert_datasets(conn, records)

        print("\n导入数据元...")
        elements = upsert_data_elements(conn, records, dataset_map)

        print("\n自动创建条件模型...")
        auto_create_condition_models(conn, elements, category_map, records)

        print("\n✅ 导入完成!")
        print(f"   数据集: {len(dataset_map)} 个")
        print(f"   数据元: {len(elements)} 个")
        print(f"   条件模型分类: {len(category_map)} 个")
        print(f"   条件模型: {len(elements)} 个")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ 导入失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()

if __name__ == '__main__':
    import json
    main()
